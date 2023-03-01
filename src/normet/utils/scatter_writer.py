import openpyxl
from bluesmet.common.scatter import Scatter
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


class ScatterExcelWriter:
    """Write a scatter table to an excel file"""

    def __init__(self, scatter: Scatter, row_name: str, column_name: str):
        self.scatter = scatter
        self.row_name = row_name
        self.column_name = column_name
        self.workbook = openpyxl.Workbook()
        self.sheet: Worksheet = self.workbook.active

    def __get_color(self, occurence: int, total: int) -> str:
        """Get a color for the cell based on the occurence and the total number of occurences"""
        if occurence == 0:
            return None

        prob = occurence / total
        # scale the probability to shift it towards the red end of the spectrum to exagerate the small values
        value = min(1.0, 5 * prob)
        red = int(255 * value)
        green = int(255 * (1 - value))
        blue = 0
        chex = [f"{i:02x}" for i in [red, green, blue]]
        return "".join(chex)

    def write_occurences(self):
        """Write the occurences to the excel file"""
        upper_row = self.scatter.upper_rows()
        upper_column = self.scatter.upper_columns()
        occurences = self.scatter.occurences()
        header = (
            [f"{self.row_name}/{self.column_name}"] + upper_column.tolist() + ["Sum"]
        )
        self.sheet.append(header)
        total_sum = occurences.sum()
        for i, occ in enumerate(occurences):
            row = [upper_row[i]] + occ.tolist() + [occ.sum()]
            self.sheet.append(row)
            for j, cell in enumerate(occ):
                color = self.__get_color(cell, total_sum)
                if color:
                    self.sheet.cell(i + 2, j + 2).fill = PatternFill(
                        "solid", start_color=color
                    )

        footer = ["Sum"] + occurences.sum(axis=0).tolist() + [total_sum]
        self.sheet.append(footer)

    def append(self, row):
        """Append a row to the excel file"""
        self.sheet.append(row)

    def write_mean_of(self, name):
        """Write the scatter table to the excel file"""
        self.sheet.append([f"Mean {name}"])
        scatter_values = self.scatter.get_values(name)
        row_vals = self.scatter.upper_rows()
        col_vals = self.scatter.upper_columns()
        self.sheet.append([f"{self.row_name}/{self.column_name}"] + col_vals.tolist())
        for i, value_row in enumerate(scatter_values):
            self.sheet.append([row_vals[i]] + value_row.tolist())
            for j, _ in enumerate(value_row):
                sc: Cell = self.sheet.cell(self.sheet.max_row, j + 1)
                sc.number_format = "0.00"

    def save(self, filename):
        """Save the excel file"""
        self.workbook.save(filename)
